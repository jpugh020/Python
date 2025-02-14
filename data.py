import pandas as pd
import numpy as np
import os
from time import sleep
import getpass
from platform import system
import re
import openpyxl
from pathlib import Path
import zipfile

import sys





platform = system()
user = getpass.getuser() 
regex = r"^\s\d{2}$"
test = r"^TEST$"
clas = r"^CLAS"
docs = f"C:\\Users\\{user}\\Documents"

if platform:
    if platform == "Windows":
        path = f'C:\\Users\\{user}\\Downloads\\'
        sudo = 'C:\\'
    elif platform == "Linux":
        path = f'~/Documents'



fileName = input("Welcome to the Jenn-Tron 5000. Please enter the name of a document you would like to Reist-ify, including the filetype at the end (e.g. .csv or .xlsx): ")
filecount = 0
for root, dirs, files in os.walk(path):
    fileLen = len(files)
    if fileLen != 0:
        if fileName in files:
            print(f"Found at {root}")
            path = root + fileName
            break
        else:
            filecount += 1
            print(f"Checked {filecount} files\r", end="")
sheet = pd.read_excel(path)
sheet = sheet.dropna(axis=1, how="all")
savePath = f"C:\\Users\\{user}\\Documents\\"

start = 0
stop = 0
sheetCount = 0
dfs = {}
dropVals = []
drops = {}
info = []
year = ''
teacher = ''
sheetName = ''
for cols in sheet:
    col = sheet[cols]
    
    for i, item in enumerate(col):
        if i == len(col) - 1:
            next = col[i]
        else:
            next = col[i+1]
        if pd.notna(item):
            if col.name == "1wbgra15.p" and type(item) == str and 'Teacher' in item:
                start = i
                
            if col.name == "1wbgra15.p" and type(item) == str and re.fullmatch(regex, item) and i > start and (pd.isna(next) or i == len(col) - 1):
                stop = i
                sheetCount += 1
                temp = sheet[start:stop]
                
                for i, temp_cols in enumerate(temp):
                    temp_col = temp[temp_cols]
                    for temp_item in temp_col:
                        if type(temp_item) == str and "School Year: " in temp_item or type(temp_item) == str and "Sec: " in temp_item:
                            info.append(temp_item)
                        
                        if type(temp_item) == str and "Period: " in temp_item:
                            sheetName = temp_item[:6]
                            sheetName += temp_item[7:]
                            dfs.update({sheetName : temp})
                            
                        if type(temp_item) == str and "CLAS" in temp_item:
                            dropVals.append(temp_cols)
                        if type(temp_item) == str and "Teacher: " in temp_item and savePath == f"C:\\Users\\{user}\\Documents\\":
                            teacher = temp_item[9:]
                            savePath += teacher + ".xlsx"
                drops[sheetName] = dropVals    
                dropVals = []
                    
                
                start = 0
                stop = 0
for root, dirs, files in os.walk(docs):
    fileLen = len(files)
    if fileLen != 0:
        if teacher + ".xlsx" in files:
            book = openpyxl.load_workbook(savePath)
            break
        else:
            book = openpyxl.Workbook()
            book.save(savePath)
            

                    
                    




writer = pd.ExcelWriter(savePath, engine='openpyxl', mode='a')
try:
    writer.workbook = book
except zipfile.BadZipFile:
    print("Bad")

for index, (key, value) in enumerate(dfs.items()):
    value = value.drop(columns=drops.get(key))
    value.to_excel(writer, index=False, header=False, sheet_name=key)

writer.close()

book = openpyxl.load_workbook(savePath)
del book['Sheet']
book.save(savePath)











students = []



# print(sheet.notna())
# print(int(col[8]))
# print(data)